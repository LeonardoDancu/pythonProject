# https://www.projectpro.io/recipes/convert-excel-document-xml-format
# https://gist.github.com/omairaasim/9de079f490b66e585d990142b7f6ab6c
from openpyxl import load_workbook
from yattag import Doc, indent

# Load our Excel File
wb = load_workbook(
    "f_martie_2023_5.xlsx",
    data_only=True
)
# https://stackoverflow.com/questions/28517508/read-excel-cell-value-and-not-the-formula-computing-it-openpyxl

# Getting an object of active sheet 1
ws = wb.worksheets[0]

# Returning returns a triplet
doc, tag, text = Doc().tagtext()

#xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
#xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'
xml_header = ''
xml_schema = ''

# Appends the String to document
doc.asis(xml_header)
doc.asis(xml_schema)

with tag('Facturi'):
    for row in ws.iter_rows(min_row=2, max_row=479, min_col=1, max_col=48):
        row = [cell.value for cell in row]
        with tag("Factura"):
            with tag("Antet"):
                with tag("FurnizorNume"):
                    if row[0] is None:
                        row[0] = ""
                        text(row[0])
                    else:
                        text(row[0])
                with tag("FurnizorCIF"):
                    if row[1] is None:
                        row[1] = ""
                        text(row[1])
                    else:
                        text(row[1])
                with tag("FurnizorNrRegCom"):
                    if row[2] is None:
                        row[2] = ""
                        text(row[2])
                    else:
                        text(row[2])
                with tag("FurnizorCapital"):
                    if row[3] is None:
                        row[3] = ""
                        text(row[3])
                    else:
                        text(row[3])
                with tag("FurnizorTara"):
                    if row[4] is None:
                        row[4] = ""
                        text(row[4])
                    else:
                        text(row[4])
                with tag("FurnizorLocalitate"):
                    if row[5] is None:
                        row[5] = ""
                        text(row[5])
                    else:
                        text(row[5])
                with tag("FurnizorJudet"):
                    if row[6] is None:
                        row[6] = ""
                        text(row[6])
                    else:
                        text(row[6])
                with tag("FurnizorAdresa"):
                    if row[7] is None:
                        row[7] = ""
                        text(row[7])
                    else:
                        text(row[7])
                with tag("FurnizorTelefon"):
                    if row[8] is None:
                        row[8] = ""
                        text(row[8])
                    else:
                        text(row[8])
                with tag("FurnizorMail"):
                    if row[9] is None:
                        row[9] = ""
                        text(row[9])
                    else:
                        text(row[9])
                with tag("FurnizorBanca"):
                    if row[10] is None:
                        row[10] = ""
                        text(row[10])
                    else:
                        text(row[10])
                with tag("FurnizorIBAN"):
                    if row[11] is None:
                        row[11] = ""
                        text(row[11])
                    else:
                        text(row[11])
                with tag("FurnizorInformatiiSuplimentare"):
                    with tag("GUID_cod_client"):
                        if row[12] is None:
                            row[12] = ""
                            text(row[12])
                        else:
                            text(row[12])
                    with tag("ClientNume"):
                        if row[13] is None:
                            row[13] = ""
                            text(row[13])
                        else:
                            text(row[13])
                with tag("ClientInformatiiSuplimentare"):
                    if row[14] is None:
                        row[14] = ""
                        text(row[14])
                    else:
                        text(row[14])
                with tag("ClientCIF"):
                    if row[15] is None:
                        row[15] = ""
                        text(row[15])
                    else:
                        text(row[15])
                with tag("ClientNrRegCom"):
                    if row[16] is None:
                        row[16] = ""
                        text(row[16])
                    else:
                        text(row[16])
                with tag("ClientJudet"):
                    if row[17] is None:
                        row[17] = ""
                        text(row[17])
                    else:
                        text(row[17])
                with tag("ClientTara"):
                    if row[18] is None:
                        row[18] = ""
                        text(row[18])
                    else:
                        text(row[18])
                with tag("ClientLocalitate"):
                    if row[19] is None:
                        row[19] = ""
                        text(row[19])
                    else:
                        text(row[19])
                with tag("ClientAdresa"):
                    if row[20] is None:
                        row[20] = ""
                        text(row[20])
                    else:
                        text(row[20])
                with tag("ClientBanca"):
                    if row[21] is None:
                        row[21] = ""
                        text(row[21])
                    else:
                        text(row[21])
                with tag("ClientIBAN"):
                    if row[22] is None:
                        row[22] = ""
                        text(row[22])
                    else:
                        text(row[22])
                with tag("ClientTelefon"):
                    if row[23] is None:
                        row[23] = ""
                        text(row[23])
                    else:
                        text(row[23])
                with tag("ClientMail"):
                    if row[24] is None:
                        row[24] = ""
                        text(row[24])
                    else:
                        text(row[24])
                with tag("FacturaNumar"):
                    if row[25] is None:
                        row[25] = ""
                        text(row[25])
                    else:
                        text(row[25])
                with tag("FacturaData"):
                    if row[26] is None:
                        row[26] = ""
                        text(row[26])
                    else:
                        text(row[26])
                with tag("FacturaScadenta"):
                    if row[27] is None:
                        row[27] = ""
                        text(row[27])
                    else:
                        text(row[27])
                with tag("FacturaTaxareInversa"):
                    if row[28] is None:
                        row[28] = ""
                        text(row[28])
                    else:
                        text(row[28])
                with tag("FacturaTVAIncasare"):
                    if row[29] is None:
                        row[29] = ""
                        text(row[29])
                    else:
                        text(row[29])
                with tag("FacturaTip"):
                    if row[30] is None:
                        row[30] = ""
                        text(row[30])
                    else:
                        text(row[30])
                with tag("FacturaInformatiiSuplimentare"):
                    print("# ce fac cu coloana 31 ce apartine chiar de FacturaInformatiiSuplimentare")
                    if row[31] is None:
                        row[31] = ""
                        text(row[31])
                    else:
                        text(row[31])
                    with tag("FacturaMoneda"):
                        if row[32] is None:
                            row[32] = ""
                            text(row[32])
                        else:
                            text(row[32])
                with tag("FacturaGreutate"):
                    if row[33] is None:
                        row[33] = ""
                        text(row[33])
                    else:
                        text(row[33])
                with tag("FacturaAccize"):
                    if row[34] is None:
                        row[34] = ""
                        text(row[34])
                    else:
                        text(row[34])
                with tag("Cod"):
                    print("# aici nu inteleg ce valoare din xlxs sa aiba tag Cod din structura")
                with tag("GUID_cod_client"):
                    print("# aici nu inteleg ce valoare din xlxs sa aiba tag GUID_cod_client din structura")
            with tag("Detalii"):
                with tag("Continut"):
                    with tag("Linie"):
                        with tag("LinieNrCrt"):
                            if row[35] is None:
                                row[35] = ""
                                text(row[35])
                            else:
                                text(row[35])
                        with tag("Gestiune"):
                            print("# aici nu inteleg ce valoare din xlxs sa aiba tag Gestiune din structura")
                        with tag("Activitate"):
                            print("# aici nu inteleg ce valoare din xlxs sa aiba tag Activitate din structura")
                        with tag("Descriere"):
                            if row[36] is None:
                                row[36] = ""
                                text(row[36])
                            else:
                                text(row[36])
                        with tag("CodArticolFurnizor"):
                            if row[37] is None:
                                row[37] = ""
                                text(row[37])
                            else:
                                text(row[37])
                        with tag("CodArticolClient"):
                            if row[38] is None:
                                row[38] = ""
                                text(row[38])
                            else:
                                text(row[38])
                            with tag("GUID_cod_articol"):
                                print("# aici nu inteleg ce valoare din xlxs sa aiba tag GUID_cod_articol din structura")
                                with tag("CodBare"):
                                    if row[39] is None:
                                        row[39] = ""
                                        text(row[39])
                                    else:
                                        text(row[39])
                        with tag("InformatiiSuplimentare"):
                            if row[40] is None:
                                row[40] = ""
                                text(row[40])
                            else:
                                text(row[40])
                        with tag("UM"):
                            if row[41] is None:
                                row[41] = ""
                                text(row[41])
                            else:
                                text(row[41])
                        with tag("Cantitate"):
                            if row[42] is None:
                                row[42] = ""
                                text(row[42])
                            else:
                                text(row[42])
                        with tag("Pret"):
                            if row[43] is None:
                                row[43] = ""
                                text(row[43])
                            else:
                                text(row[43])
                        with tag("Valoare"):
                            if row[44] is None:
                                row[44] = ""
                                text(row[44])
                            else:
                                text(row[44])
                        with tag("ProcTVA"):
                            if row[45] is None:
                                row[45] = ""
                                text(row[45])
                            else:
                                text(row[45])
                        with tag("TVA"):
                            if row[46] is None:
                                row[46] = ""
                                text(row[46])
                            else:
                                text(row[46])
                                print((row[46]))
                        with tag("Cont"):
                            if row[47] is None:
                                row[47] = ""
                                text(row[47])
                            else:
                                text(row[47])

result = indent(
    doc.getvalue(),
    indentation=' ',
    indent_text=True
)

#print(result)
with open("F_cod-fiscal_numar-factura_data-factura.xml", "w") as f:
    f.write(result)